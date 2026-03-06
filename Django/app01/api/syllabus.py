import random
import re
from rest_framework.decorators import api_view
from rest_framework.response import Response
from ..models import Question, PracticeAnswer
from .utils import require_login
from .utils import enforce_face_required  # 新增：人脸访问门槛
# 新增：考纲条目模型
from ..models import SyllabusItem
from django.http import HttpResponse
from django.db import IntegrityError
import openpyxl
from openpyxl import Workbook

# 考纲/知识点练习：支持按题型/难度/知识点筛选，以及补弱模式、全覆盖模式。
# knowledge_points 字段逗号分隔；前端传参 knowledge（或 kp） 指定单个知识点。

@api_view(["GET"])  # 练习可选项（题型/难度/知识点）
def api_syllabus_options(request):
    user, err = require_login(request, roles=['学生','VIP'])
    if err: return err
    if user.role in ('学生','VIP'):
        gate = enforce_face_required(request)
        if gate: return gate
    qs = Question.objects.filter(question_type__in=('单选','多选','判断'))
    type_counts = {}
    difficulty_counts = {}
    difficulties = set()
    kp_set = set()
    primary_set = set()
    # 新增：kp -> primary 映射
    kp_to_primary = {}
    for q in qs:
        type_counts[q.question_type] = type_counts.get(q.question_type, 0) + 1
        difficulties.add(q.score)
        difficulty_counts[q.score] = difficulty_counts.get(q.score, 0) + 1
        if q.knowledge_points:
            raw = q.knowledge_points.replace('，', ',')
            kps = []
            for part in raw.split(','):
                p = part.strip()
                if p:
                    kp_set.add(p)
                    kps.append(p)
            # 记录映射
            if q.primary_knowledge:
                pk = q.primary_knowledge.strip()
                if pk:
                    for kp in kps:
                        s = kp_to_primary.setdefault(kp, set())
                        s.add(pk)
        if q.primary_knowledge:
            p = q.primary_knowledge.strip()
            if p:
                primary_set.add(p)
    types = [t for t in ['单选','多选','判断'] if type_counts.get(t)]
    # 将映射的集合转为有序列表
    kp_to_primary_out = {k: sorted(list(v)) for k, v in kp_to_primary.items()}
    return Response({
        'success': True,
        'types': types,
        'difficulties': sorted(list(difficulties)),
        'knowledge_points': sorted(list(kp_set)),
        'primary_points': sorted(list(primary_set)),  # 新增：一级知识点集合
        'kp_to_primary': kp_to_primary_out,           # 新增：联动映射
        'counts_by_type': {k:int(v) for k,v in type_counts.items()},
        'counts_by_difficulty': {int(k):int(v) for k,v in difficulty_counts.items()},
        'total': sum(type_counts.values())
    })

@api_view(["GET"])  # 知识点掌握度统计（按用户练习记录汇总）
def api_syllabus_stats(request):
    user, err = require_login(request, roles=['学生','VIP'])
    if err: return err
    if user.role in ('学生','VIP'):
        gate = enforce_face_required(request)
        if gate: return gate
    uid = user.user_id
    level = request.query_params.get('level') or 'knowledge'
    level = 'primary' if str(level).lower().startswith('p') else 'knowledge'
    kp_map = {}
    if not uid:
        return Response({'success': True, 'kp_stats': [], 'level': level})
    pas = PracticeAnswer.objects.filter(user_id=uid)
    if not pas.exists():
        return Response({'success': True, 'kp_stats': [], 'level': level})
    qids = [p.question_id for p in pas]
    qmap = {q.question_id: q for q in Question.objects.filter(question_id__in=qids)}
    for p in pas:
        q = qmap.get(p.question_id)
        if not q:
            continue
        if level == 'knowledge':
            if q.knowledge_points:
                raw = q.knowledge_points.replace('，', ',')
                for part in raw.split(','):
                    name = part.strip()
                    if not name:
                        continue
                    stat = kp_map.setdefault(name, {'name': name, 'total_attempts': 0, 'correct_attempts': 0})
                    stat['total_attempts'] += p.total_attempts
                    stat['correct_attempts'] += p.correct_attempts
        else:  # primary
            if q.primary_knowledge:
                name = q.primary_knowledge.strip()
                if name:
                    stat = kp_map.setdefault(name, {'name': name, 'total_attempts': 0, 'correct_attempts': 0})
                    stat['total_attempts'] += p.total_attempts
                    stat['correct_attempts'] += p.correct_attempts
    out = []
    for name, stat in kp_map.items():
        tot = stat['total_attempts']
        cor = stat['correct_attempts']
        acc = round(100.0 * cor / tot, 1) if tot else None
        out.append({'name': name, 'total_attempts': tot, 'correct_attempts': cor, 'accuracy': acc})
    out.sort(key=lambda x: (9999 if x['accuracy'] is None else x['accuracy']))
    return Response({'success': True, 'kp_stats': out, 'level': level})

@api_view(["GET"])  # 按条件/模式抽题：mode=normal|weak|cover
def api_syllabus_questions(request):
    user, err = require_login(request, roles=['学生','VIP'])
    if err: return err
    if user.role in ('学生','VIP'):
        gate = enforce_face_required(request)
        if gate: return gate
    uid = user.user_id

    mode = (request.query_params.get('mode') or 'normal').strip().lower()
    level = request.query_params.get('level') or request.query_params.get('kp_level') or 'knowledge'
    level = 'primary' if str(level).lower().startswith('p') else 'knowledge'
    qtype = request.query_params.get('type') or ''
    try:
        difficulty = request.query_params.get('difficulty')
        difficulty = int(difficulty) if difficulty not in (None,'') else None
    except Exception:
        difficulty = None
    knowledge = request.query_params.get('knowledge') or request.query_params.get('kp') or ''
    knowledge = knowledge.strip()
    try:
        limit = int(request.query_params.get('limit') or 10)
    except Exception:
        limit = 10
    limit = max(1, min(300, limit))
    shuffle = request.query_params.get('shuffle')
    do_shuffle = str(shuffle) in {'1','true','True','yes','on'}

    # cover 模式：per_kp 每知识点题数；kp_list 指定知识点子集
    try:
        per_kp = int(request.query_params.get('per_kp') or 1)
    except Exception:
        per_kp = 1
    per_kp = max(1, min(20, per_kp))
    kp_list_param = request.query_params.get('kp_list') or ''
    kp_list = [s.strip() for s in kp_list_param.split(',') if s.strip()] if kp_list_param else []

    # weak 模式：优先薄弱知识点（准确率低）
    uid = request.session.get('user_id')

    base_qs = Question.objects.filter(question_type__in=('单选','多选','判断'))
    if qtype:
        base_qs = base_qs.filter(question_type=qtype)
    if difficulty is not None:
        base_qs = base_qs.filter(score=difficulty)
    if knowledge and mode == 'normal':  # normal 模式可用单知识点过滤
        if level == 'knowledge':
            base_qs = base_qs.filter(knowledge_points__icontains=knowledge)
        else:
            base_qs = base_qs.filter(primary_knowledge__icontains=knowledge)

    def build_out(picked):
        out = []
        for q in picked:
            options = []
            if q.question_type in ['单选','多选']:
                for key,label in zip(['A','B','C','D'], [q.A_answer, q.B_answer, q.C_answer, q.D_answer]):
                    if label:
                        options.append({'key': key, 'label': label})
            elif q.question_type == '判断':
                options = [{'key':'A','label':'正确'},{'key':'B','label':'错误'}]
            out.append({
                'id': q.question_id,
                'type': q.question_type,
                'content': q.content,
                'options': options,
                'score': q.score,
                'analysis': q.analysis or '',
                'knowledge_points': q.knowledge_points or '',
                'primary_knowledge': q.primary_knowledge or ''
            })
        return out

    # 构建知识点 -> 题目映射（根据 level 决定）
    all_questions = list(base_qs)
    if not all_questions and mode in {'cover','weak'}:
        return Response({'success': True, 'questions': []})
    kp_to_questions = {}
    if mode in {'cover','weak'}:
        for q in all_questions:
            if level == 'knowledge':
                if q.knowledge_points:
                    raw = q.knowledge_points.replace('，', ',')
                    seen = set()
                    for part in raw.split(','):
                        name = part.strip()
                        if not name or name in seen:
                            continue
                        seen.add(name)
                        kp_to_questions.setdefault(name, []).append(q)
            else:
                if q.primary_knowledge:
                    name = q.primary_knowledge.strip()
                    if name:
                        kp_to_questions.setdefault(name, []).append(q)

    if mode == 'cover':
        target_list_param = request.query_params.get('kp_list') or ''
        target_list = [s.strip() for s in target_list_param.split(',') if s.strip()] if target_list_param else []
        target_kps = target_list if target_list else list(kp_to_questions.keys())
        picked = []
        for kp in target_kps:
            qlist = kp_to_questions.get(kp, [])
            if not qlist:
                continue
            sample = qlist if len(qlist) <= per_kp else random.sample(qlist, per_kp)
            picked.extend(sample)
        # 去重（按题目 ID）
        uniq = {}
        for q in picked:
            uniq[q.question_id] = q
        final = list(uniq.values())
        if do_shuffle:
            random.shuffle(final)
        if limit and len(final) > limit:
            final = final[:limit]
        return Response({'success': True, 'questions': build_out(final), 'level': level})

    if mode == 'weak':
        if not uid:
            return Response({'success': True, 'questions': [], 'error_msg': '未登录，无法补弱', 'level': level})
        # 计算知识点准确率（按当前 level）
        pas = PracticeAnswer.objects.filter(user_id=uid)
        kp_stats = {}
        if pas.exists():
            qids = [p.question_id for p in pas]
            qmap = {q.question_id: q for q in Question.objects.filter(question_id__in=qids)}
            for p in pas:
                q = qmap.get(p.question_id)
                if not q:
                    continue
                if level == 'knowledge':
                    if not q.knowledge_points:
                        continue
                    raw = q.knowledge_points.replace('，', ',')
                    for part in raw.split(','):
                        name = part.strip()
                        if not name:
                            continue
                        stat = kp_stats.setdefault(name, {'total': 0, 'correct': 0})
                        stat['total'] += p.total_attempts
                        stat['correct'] += p.correct_attempts
                else:  # primary
                    if not q.primary_knowledge:
                        continue
                    name = q.primary_knowledge.strip()
                    if not name:
                        continue
                    stat = kp_stats.setdefault(name, {'total': 0, 'correct': 0})
                    stat['total'] += p.total_attempts
                    stat['correct'] += p.correct_attempts
        # 生成按准确率升序的知识点列表（未练过的 accuracy 视为 0 优先）
        scored_kps = []
        for kp, qlist in kp_to_questions.items():
            s = kp_stats.get(kp)
            if not s or s['total'] == 0:
                acc = 0.0
            else:
                acc = 100.0 * s['correct'] / s['total'] if s['total'] else 0.0
            scored_kps.append((acc, kp))
        scored_kps.sort(key=lambda x: x[0])  # 低准确率在前
        picked = []
        for _, kp in scored_kps:
            plist = kp_to_questions.get(kp, [])
            random.shuffle(plist)
            for q in plist:
                if q not in picked:
                    picked.append(q)
                if len(picked) >= limit:
                    break
            if len(picked) >= limit:
                break
        return Response({'success': True, 'questions': build_out(picked), 'level': level})

    # normal 分支保留原逻辑
    if mode == 'normal':
        qlist = list(base_qs)
        if not qlist:
            return Response({'success': True, 'questions': [], 'level': level})
        if len(qlist) <= limit:
            picked = qlist
            if do_shuffle:
                random.shuffle(picked)
        else:
            picked = random.sample(qlist, limit)
        return Response({'success': True, 'questions': build_out(picked), 'level': level})

    # 未识别 mode
    return Response({'success': True, 'questions': []})


# ========== 新增：考纲预置导入与获取 ==========

@api_view(["POST"])  # 管理端：考纲一键导入（文本）
def api_manage_syllabus_import_text(request):
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    province = (request.data.get('province') or '').strip()
    major = (request.data.get('major') or '').strip()
    text = request.data.get('text')
    mode = (request.data.get('mode') or 'replace').strip().lower()  # replace/append
    if not province or not major or not text:
        return Response({'success': False, 'error_msg': '省份/专业/文本 不能为空'})
    if mode not in {'replace','append'}:
        mode = 'replace'
    # 清空旧数据
    if mode == 'replace':
        SyllabusItem.objects.filter(province=province, major=major).delete()
    lines = str(text).splitlines()
    cur_kp = None
    created = 0
    skipped = 0
    kp_count = 0
    for raw in lines:
        s = str(raw or '').strip()
        if not s:
            continue
        m_kp = re.match(r'^【(.+?)】$', s)
        if m_kp:
            cur_kp = m_kp.group(1).strip()
            if cur_kp:
                kp_count += 1
            continue
        m_item = re.match(r'^(\d+)[.、]\s*(.+)$', s)
        if m_item and cur_kp:
            primary = m_item.group(2).strip()
            if not primary:
                continue
            try:
                SyllabusItem.objects.create(province=province, major=major, kp=cur_kp, primary=primary)
                created += 1
            except IntegrityError:
                skipped += 1
            except Exception:
                pass
    return Response({'success': True, 'created': created, 'skipped': skipped, 'kp_sections': kp_count, 'mode': mode})


@api_view(["GET"])  # 获取考纲预置（树/指定省份专业的映射）
def api_syllabus_presets(request):
    # 无参：返回 {provinces, majors_by_province}
    province = (request.query_params.get('province') or '').strip()
    major = (request.query_params.get('major') or '').strip()
    if not province or not major:
        provs = sorted(set(SyllabusItem.objects.values_list('province', flat=True)))
        m_map = {}
        for p in provs:
            majors = sorted(set(SyllabusItem.objects.filter(province=p).values_list('major', flat=True)))
            m_map[p] = majors
        return Response({'success': True, 'provinces': provs, 'majors_by_province': m_map})
    # 指定省市专业：返回 kp->primary
    rows = SyllabusItem.objects.filter(province=province, major=major)
    kp_map = {}
    prim_set = set()
    for it in rows:
        kp_map.setdefault(it.kp, []).append(it.primary)
        prim_set.add(it.primary)
    # 排序
    kp_to_primary = {k: sorted(v) for k, v in kp_map.items()}
    kp_list = sorted(kp_to_primary.keys())
    return Response({'success': True,
                     'province': province,
                     'major': major,
                     'kp_list': kp_list,
                     'primary_list': sorted(list(prim_set)),
                     'kp_to_primary': kp_to_primary})


# ========== 管理端：考纲 CRUD ==========

@api_view(["GET"])  # 管理端：列表（按省份/专业，可筛选kp/关键字，分页）
def api_manage_syllabus_list(request):
    # 复用排序规则
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    province = (request.query_params.get('province') or '').strip()
    major = (request.query_params.get('major') or '').strip()
    if not province or not major:
        provs = sorted(set(SyllabusItem.objects.values_list('province', flat=True)))
        majors_by = {}
        for p in provs:
            majors_by[p] = sorted(set(SyllabusItem.objects.filter(province=p).values_list('major', flat=True)))
        return Response({'success': True, 'items': [], 'total': 0, 'provinces': provs, 'majors_by_province': majors_by})
    kp = (request.query_params.get('kp') or '').strip()
    kw = (request.query_params.get('q') or request.query_params.get('keyword') or '').strip()
    try:
        page = int(request.query_params.get('page') or 1)
    except Exception:
        page = 1
    try:
        page_size = int(request.query_params.get('page_size') or 20)
    except Exception:
        page_size = 20
    page_size = max(1, min(200, page_size))
    q = _ordered_qs(province, major, kp if kp else None, kw if kw else None)
    total = q.count()
    start = (page - 1) * page_size
    rows = q[start:start+page_size]
    items = [{'id': r.id, 'kp': r.kp, 'primary': r.primary} for r in rows]
    # 聚合：kp 列表与数量
    kp_counts = {}
    for r in SyllabusItem.objects.filter(province=province, major=major).values('kp'):
        k = r['kp']
        kp_counts[k] = kp_counts.get(k, 0) + 1
    kp_list = sorted(kp_counts.keys())
    return Response({'success': True, 'items': items, 'total': total, 'kp_list': kp_list, 'kp_counts': kp_counts, 'province': province, 'major': major})

@api_view(["POST"])  # 管理端：保存（新增/编辑）
def api_manage_syllabus_save(request):
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    try:
        iid = request.data.get('id')
        province = (request.data.get('province') or '').strip()
        major = (request.data.get('major') or '').strip()
        kp = (request.data.get('kp') or '').strip()
        primary = (request.data.get('primary') or '').strip()
    except Exception:
        return Response({'success': False, 'error_msg': '参数错误'})
    if not province or not major or not kp or not primary:
        return Response({'success': False, 'error_msg': '省份/专业/大类/一级知识点不能为空'})
    if iid:
        try:
            obj = SyllabusItem.objects.get(id=int(iid))
        except Exception:
            return Response({'success': False, 'error_msg': '记录不存在'})
        obj.province = province
        obj.major = major
        obj.kp = kp
        obj.primary = primary
        obj.save()
        return Response({'success': True, 'id': obj.id})
    else:
        obj = SyllabusItem.objects.create(province=province, major=major, kp=kp, primary=primary)
        return Response({'success': True, 'id': obj.id})

@api_view(["POST"])  # 管理端：删除（单条或按条件批量）
def api_manage_syllabus_delete(request):
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    iid = request.data.get('id')
    if iid:
        try:
            SyllabusItem.objects.get(id=int(iid)).delete()
            return Response({'success': True, 'deleted': 1})
        except Exception:
            return Response({'success': False, 'error_msg': '记录不存在'})
    # 批量按 province/major/kp 删除
    province = (request.data.get('province') or '').strip()
    major = (request.data.get('major') or '').strip()
    kp = (request.data.get('kp') or '').strip()
    q = SyllabusItem.objects.all()
    if province: q = q.filter(province=province)
    if major: q = q.filter(major=major)
    if kp: q = q.filter(kp=kp)
    n = q.count()
    if n:
        q.delete()
    return Response({'success': True, 'deleted': n})

@api_view(["POST"])  # 管理端：清空某省份/专业（或某大类）
def api_manage_syllabus_clear(request):
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    province = (request.data.get('province') or '').strip()
    major = (request.data.get('major') or '').strip()
    kp = (request.data.get('kp') or '').strip()
    if not province or not major:
        return Response({'success': False, 'error_msg': '省份/专业不能为空'})
    q = SyllabusItem.objects.filter(province=province, major=major)
    if kp:
        q = q.filter(kp=kp)
    n = q.count()
    if n:
        q.delete()
    return Response({'success': True, 'deleted': n})

@api_view(["GET"])  # 管理端：导出（fmt=excel|md），可选 kp 过滤
def api_manage_syllabus_export(request):
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    province = (request.query_params.get('province') or '').strip()
    major = (request.query_params.get('major') or '').strip()
    fmt = (request.query_params.get('fmt') or 'excel').strip().lower()
    kp = (request.query_params.get('kp') or '').strip()
    if not province or not major:
        return Response({'success': False, 'error_msg': '省份/专业不能为空'})
    q = SyllabusItem.objects.filter(province=province, major=major)
    if kp:
        q = q.filter(kp=kp)
    rows = list(q.order_by('kp','primary','id'))
    if fmt == 'md' or fmt == 'markdown':
        # 生成 Markdown
        lines = [f"# {province}·{major} 考纲"]
        cur = None
        idx = 0
        for r in rows:
            if r.kp != cur:
                if cur is not None:
                    lines.append('')
                cur = r.kp
                lines.append(f"\n【{cur}】")
                idx = 0
            idx += 1
            lines.append(f"{idx}. {r.primary}")
        content = "\n".join(lines)
        resp = HttpResponse(content, content_type='text/markdown; charset=utf-8')
        resp['Content-Disposition'] = f"attachment; filename=syllabus_{province}_{major}.md"
        return resp
    # 默认 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'syllabus'
    ws.append(['省份','专业','考纲大类','一级知识点'])
    for r in rows:
        ws.append([r.province, r.major, r.kp, r.primary])
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename=syllabus_{province}_{major}.xlsx"
    return resp

@api_view(["GET"])  # 管理端：导入模板（markdown）
def api_manage_syllabus_template(request):
    sample = (
        "【计算机基础】\n"
        "1. 了解计算机的发展、特点、分类及应用领域\n"
        "2. 了解计算机的工作原理，熟悉计算机系统的组成\n\n"
        "【数据库基础】\n"
        "1. 了解数据、数据库、数据库系统及数据库管理系统等概念\n"
        "2. 理解实体模型的相关术语以及实体间的关系\n"
    )
    resp = HttpResponse(sample, content_type='text/plain; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename=syllabus_template.md'
    return resp

@api_view(["POST"])  # 管理端：去重（按省份/专业，保留较小ID）
def api_manage_syllabus_dedupe(request):
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    province = (request.data.get('province') or '').strip()
    major = (request.data.get('major') or '').strip()
    dry = str(request.data.get('dry') or '0').lower() in ('1','true','yes')
    if not province or not major:
        return Response({'success': False, 'error_msg': '省份/专业不能为空'})
    q = SyllabusItem.objects.filter(province=province, major=major).order_by('kp','primary','id')
    seen = {}
    to_delete = []
    for r in q:
        key = (r.kp, r.primary)
        if key in seen:
            to_delete.append(r.id)
        else:
            seen[key] = r.id
    deleted = 0
    if to_delete and not dry:
        SyllabusItem.objects.filter(id__in=to_delete).delete()
        deleted = len(to_delete)
    return Response({'success': True, 'duplicates': len(to_delete), 'deleted': deleted, 'dry_run': dry})

@api_view(["POST"])  # 管理端：大类重命名（级联，可能触发合并）
def api_manage_syllabus_rename_kp(request):
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    province = (request.data.get('province') or '').strip()
    major = (request.data.get('major') or '').strip()
    from_kp = (request.data.get('from_kp') or '').strip()
    to_kp = (request.data.get('to_kp') or '').strip()
    if not (province and major and from_kp and to_kp):
        return Response({'success': False, 'error_msg': '省份/专业/原大类/新大类 均不能为空'})
    if from_kp == to_kp:
        return Response({'success': True, 'updated': 0, 'merged': 0})
    qs = list(SyllabusItem.objects.filter(province=province, major=major, kp=from_kp).order_by('id'))
    updated = 0
    merged = 0
    for r in qs:
        try:
            r.kp = to_kp
            r.save(update_fields=['kp'])
            updated += 1
        except IntegrityError:
            # 已存在目标 (province,major,to_kp,primary) -> 合并：删除当前
            try:
                SyllabusItem.objects.filter(id=r.id).delete()
                merged += 1
            except Exception:
                pass
    return Response({'success': True, 'updated': updated, 'merged': merged})

# 调整列表排序：按大类与条目文字
def _ordered_qs(province, major, kp=None, keyword=None, page=1, page_size=20):
    q = SyllabusItem.objects.filter(province=province, major=major)
    if kp:
        q = q.filter(kp__icontains=kp)
    if keyword:
        q = q.filter(primary__icontains=keyword)
    return q.order_by('kp','primary','id')


@api_view(["POST"])  # 管理端：考纲 Excel 批量导入（multipart）
def api_manage_syllabus_import_excel(request):
    _, err = require_login(request, roles=['老师','管理员'])
    if err: return err
    file = request.FILES.get('file') or request.FILES.get('excel') or request.FILES.get('excel_file')
    if not file:
        return Response({'success': False, 'error_msg': '未上传 Excel 文件（表单字段名 file/excel/excel_file 均可）'})
    mode = (request.POST.get('mode') or request.data.get('mode') or 'append').strip().lower()
    fixed_prov = (request.POST.get('province') or request.data.get('province') or '').strip()
    fixed_major = (request.POST.get('major') or request.data.get('major') or '').strip()
    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
    except Exception:
        return Response({'success': False, 'error_msg': 'Excel 文件无法读取'})
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return Response({'success': False, 'error_msg': 'Excel 内容为空'})
    header = [str(c or '').strip() for c in rows[0]]
    def find_idx(names):
        for i, h in enumerate(header):
            if h in names: return i
        return -1
    idx_prov = find_idx({'省份','province','Province'})
    idx_major = find_idx({'专业','major','Major'})
    idx_kp = find_idx({'考纲大类','大类','kp','KP','Kp'})
    idx_primary = find_idx({'一级知识点','primary','Primary'})
    if idx_kp < 0 or idx_primary < 0:
        return Response({'success': False, 'error_msg': '缺少必要列：考纲大类/一级知识点'})
    if (idx_prov < 0 or idx_major < 0) and not (fixed_prov and fixed_major):
        return Response({'success': False, 'error_msg': '缺少 省份/专业 列，或在参数中指定 province/major'})
    data_rows = rows[1:]
    # 预扫描 distinct (prov, major) 以便 replace
    pairs = set()
    for r in data_rows:
        kp = (str(r[idx_kp]) if idx_kp >=0 else '').strip() if r and len(r)>max(idx_kp,0) else ''
        primary = (str(r[idx_primary]) if idx_primary>=0 else '').strip() if r and len(r)>max(idx_primary,0) else ''
        if not kp or not primary:
            continue
        prov = (str(r[idx_prov]) if idx_prov>=0 else '').strip() if idx_prov>=0 and r and len(r)>idx_prov else fixed_prov
        major = (str(r[idx_major]) if idx_major>=0 else '').strip() if idx_major>=0 and r and len(r)>idx_major else fixed_major
        if prov and major:
            pairs.add((prov, major))
    if mode == 'replace' and pairs:
        for prov, maj in pairs:
            SyllabusItem.objects.filter(province=prov, major=maj).delete()
    created = 0
    skipped = 0
    errors = 0
    error_rows = []
    for ridx, r in enumerate(data_rows, start=2):
        try:
            kp = (str(r[idx_kp]) if idx_kp>=0 else '').strip()
            primary = (str(r[idx_primary]) if idx_primary>=0 else '').strip()
            prov = (str(r[idx_prov]) if idx_prov>=0 else '').strip() if idx_prov>=0 else fixed_prov
            maj = (str(r[idx_major]) if idx_major>=0 else '').strip() if idx_major>=0 else fixed_major
            if not kp or not primary or not prov or not maj:
                errors += 1
                error_rows.append(ridx)
                continue
            try:
                SyllabusItem.objects.create(province=prov, major=maj, kp=kp, primary=primary)
                created += 1
            except IntegrityError:
                skipped += 1
            except Exception:
                errors += 1
                error_rows.append(ridx)
        except Exception:
            errors += 1
            error_rows.append(ridx)
    return Response({'success': True, 'created': created, 'skipped': skipped, 'errors': errors, 'error_rows': error_rows[:100], 'mode': mode, 'pairs': len(pairs)})

@api_view(["GET"])  # 管理端：Excel 导入模板（仅表头+示例）
def api_manage_syllabus_template_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'syllabus'
    ws.append(['省份','专业','考纲大类','一级知识点'])
    ws.append(['四川省','计算机类','计算机基础','了解计算机的发展、特点、分类及应用领域'])
    ws.append(['四川省','计算机类','数据库基础','了解数据、数据库、数据库系统及数据库管理系统等概念'])
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename=syllabus_import_template.xlsx'
    return resp

@api_view(["GET"])  # 当前会话的考纲选择（省份/专业）
def api_syllabus_selection(request):
    user, err = require_login(request, roles=['学生','VIP','老师','管理员'])
    if err: return err
    prov = request.session.get('syllabus_province') or ''
    major = request.session.get('syllabus_major') or ''
    return Response({'success': True, 'province': prov, 'major': major})

@api_view(["POST"])  # 保存当前会话的考纲选择（省份/专业）
def api_syllabus_selection_save(request):
    user, err = require_login(request, roles=['学生','VIP','老师','管理员'])
    if err: return err
    prov = (request.data.get('province') or '').strip()
    major = (request.data.get('major') or '').strip()
    # 允许为空（表示不限定考纲）
    request.session['syllabus_province'] = prov
    request.session['syllabus_major'] = major if prov else ''
    return Response({'success': True, 'province': prov, 'major': (major if prov else '')})
