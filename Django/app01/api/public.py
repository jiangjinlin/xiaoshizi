from datetime import timedelta
import os
import re
import tempfile as _tmp
import openpyxl
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.contrib.auth.hashers import make_password  # 新增：密码哈希
from django.db import transaction as dj_tx
from django.conf import settings
from ..models import User, ExamRecord, ExamQuestion, Question, Exam, AnswerRecord, OperationDetail
from ..scoring import score_word, score_excel, score_ppt
from .utils import fmt_dt, get_operation_template_url, norm_tf
# 新增导入：评论、登录校验
from .utils import require_login
# 新增导入：题目评论与练习记录
from ..models import QuestionComment, PracticeAnswer
# 新增导入：评论点赞与聚合
from ..models import QuestionCommentLike
from django.db.models import Count


@api_view(["GET"])  # 健康检查
def index(request):
    return Response({'success': True, 'message': '欢迎访问考试系统 API'})


@api_view(["GET"])  # 首页概览
def api_overview(request):
    from django.db import models as dj_models
    # 选择目标考试：最近一场有成绩记录的已发布考试；否则最近一场已发布考试
    target_exam = None
    try:
        recent_published = Exam.objects.filter(is_published=True).order_by('-start_time')
        for e in recent_published:
            if ExamRecord.objects.filter(exam=e).exists():
                target_exam = e
                break
        if not target_exam:
            target_exam = recent_published.first()
    except Exception:
        target_exam = None

    bins = ['60分以下', '60-70分', '70-80分', '80-90分', '90分以上']
    counts = [0, 0, 0, 0, 0]

    def bucket(p):
        if p < 60: return 0
        if p < 70: return 1
        if p < 80: return 2
        if p < 90: return 3
        return 4

    avg_score = 0.0
    pass_rate = 0.0
    exam_title = ''

    # 计算分布与均分、通过率
    if target_exam:
        exam_title = target_exam.title
        batch_ids = [eq.batch for eq in ExamQuestion.objects.filter(exam=target_exam)]
        full_score = sum(getattr(q, 'score', 1) for q in Question.objects.filter(batch__in=batch_ids)) or 0
        qs = ExamRecord.objects.filter(exam=target_exam)
        total = qs.count() or 0
        if full_score > 0:
            for r in qs.only('score'):
                p = 100.0 * float(r.score or 0) / full_score
                counts[bucket(p)] += 1
        try:
            avg_score = float(qs.aggregate(avg=dj_models.Avg('score')).get('avg') or 0)
        except Exception:
            avg_score = 0.0
        passed = 0
        if full_score > 0:
            for r in qs.only('score'):
                if (float(r.score or 0) / full_score) >= 0.6:
                    passed += 1
        pass_rate = round(100.0 * passed / total, 1) if total else 0.0

    # 题型正确率（仅统计该考试中客观题：单选/多选/判断）
    accuracy = []
    if target_exam:
        try:
            batch_ids = [eq.batch for eq in ExamQuestion.objects.filter(exam=target_exam)]
            q_map = {q.question_id: q for q in Question.objects.filter(batch__in=batch_ids)}
            ans_qs = AnswerRecord.objects.filter(record__exam=target_exam, question_id__in=q_map.keys())
            stats = {}
            for a in ans_qs.select_related('question'):
                qtype = a.question.question_type
                if qtype not in ('单选', '多选', '判断'):
                    continue
                if qtype not in stats:
                    stats[qtype] = {'ok': 0, 'total': 0}
                stats[qtype]['total'] += 1
                if bool(a.is_correct):
                    stats[qtype]['ok'] += 1
            for t in ['选择题', '多选题', '判断题']:
                key = '单选' if t == '选择题' else ('多选' if t == '多选题' else '判断')
                rec = stats.get(key, {'ok': 0, 'total': 0})
                val = round(100.0 * rec['ok'] / rec['total'], 1) if rec['total'] else 0.0
                accuracy.append({'name': t, 'value': val})
        except Exception:
            accuracy = []

    return Response({
        'success': True,
        'exam_title': exam_title,
        'score_bins': bins,
        'score_counts': counts,
        'avg_score': round(avg_score, 1),
        'pass_rate': pass_rate,
        'accuracy': accuracy,
    })

@api_view(["GET"])  # 当前有效考试及题目
def api_exam_list(request):
    now = timezone.now()
    exams = Exam.objects.filter(is_published=True, start_time__lte=now, end_time__gte=now)
    exam_data = []
    for exam in exams:
        exam_questions = ExamQuestion.objects.filter(exam=exam)
        questions = []
        for eq in exam_questions:
            for q in Question.objects.filter(batch=eq.batch):
                options = []
                if q.question_type in ['单选', '多选']:
                    for key, label in zip(['A', 'B', 'C', 'D'], [q.A_answer, q.B_answer, q.C_answer, q.D_answer]):
                        if label:
                            options.append({'key': key, 'label': label})
                elif q.question_type == '判断':
                    options = [{'key': 'A', 'label': '正确'}, {'key': 'B', 'label': '错误'}]
                data = {
                    'id': q.question_id,
                    'type': '操作' if q.question_type in ('操作', 'Office操作') else q.question_type,
                    'content': q.content,
                    'options': options,
                    'score': q.score
                }
                if q.question_type in ('操作', 'Office操作'):
                    tpl = get_operation_template_url(q, request)
                    if tpl:
                        data['template_url'] = tpl
                questions.append(data)
        exam_data.append({
            'id': exam.exam_id,
            'title': exam.title,
            'start_time': fmt_dt(exam.start_time),
            'end_time': fmt_dt(exam.end_time),
            'duration': exam.duration,
            'questions': questions
        })
    return Response({'success': True, 'exams': exam_data})


@api_view(["GET"])  # 成绩列表（个人）
def api_score_query(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return Response({'success': False, 'error_msg': '未登录'}, status=401)
    try:
        user = User.objects.get(user_id=user_id)
    except User.DoesNotExist:
        return Response({'success': False, 'error_msg': '用户不存在'}, status=401)

    delete_all = request.query_params.get('delete_all')
    delete_id = request.query_params.get('delete_id')

    if delete_all:
        ExamRecord.objects.filter(user=user).delete()
        return Response({'success': True, 'message': '已清空当前用户的成绩'})

    if delete_id:
        try:
            rec = ExamRecord.objects.get(record_id=int(delete_id), user=user)
        except (ExamRecord.DoesNotExist, ValueError):
            return Response({'success': False, 'error_msg': '记录不存在或无权删除'})
        rec.delete()
        return Response({'success': True, 'message': '删除成功'})

    records = ExamRecord.objects.filter(user=user).order_by('-submit_time')
    score_list = []
    for record in records:
        batch_ids = [eq.batch for eq in ExamQuestion.objects.filter(exam=record.exam)]
        full_score = sum(getattr(q, 'score', 1) for q in Question.objects.filter(batch__in=batch_ids))
        score_list.append({
            'exam_id': record.exam.exam_id,
            'exam_title': record.exam.title,
            'score': record.score,
            'full_score': full_score,
            'start_time': fmt_dt(record.start_time),
            'end_time': fmt_dt(record.end_time),
            'submit_time': fmt_dt(record.submit_time),
            'record_id': record.record_id
        })
    return Response({'success': True, 'score_list': score_list})


@api_view(["POST"])  # 离线评分
def api_score_word(request):
    file_path = request.data.get("file_path")
    if not file_path:
        return Response({"success": False, "error_msg": "缺少文件路径"}, status=400)
    score, total, msg = score_word(file_path)
    return Response({"success": True, "score": score, "total": total, "msg": msg})


@api_view(["POST"])  # 离线评分
def api_score_excel(request):
    file_path = request.data.get("file_path")
    if not file_path:
        return Response({"success": False, "error_msg": "缺少文件路径"}, status=400)
    score, total, msg = score_excel(file_path)
    return Response({"success": True, "score": score, "total": total, "msg": msg})


@api_view(["POST"])  # 离线评分
def api_score_ppt(request):
    file_path = request.data.get("file_path")
    if not file_path:
        return Response({"success": False, "error_msg": "缺少文件路径"}, status=400)
    score, total, msg = score_ppt(file_path)
    return Response({"success": True, "score": score, "total": total, "msg": msg})


@api_view(["POST"])  # 注册（仅开放学生）
def api_register(request):
    # 注册开关 + 角色策略说明：仅允许学生自助注册；老师/VIP 由管理员后台创建；管理员仅开发者在数据库添加
    if not getattr(settings, 'ALLOW_PUBLIC_REGISTRATION', True):
        return Response({'success': False, 'error_msg': '已关闭公开注册'}, status=403)

    username = (request.data.get('username') or '').strip()
    password = (request.data.get('password') or '').strip()
    role_req = (request.data.get('role') or '学生').strip() or '学生'
    classroom = (request.data.get('classroom') or '').strip()

    # 强制学生角色（忽略前端传来的其他角色，避免绕过）
    role = '学生'

    if not username or not password:
        return Response({'success': False, 'error_msg': '用户名/密码不能为空'}, status=400)
    if len(username) < 3:
        return Response({'success': False, 'error_msg': '用户名长度至少为3个字符'}, status=400)
    if len(password) < 6:
        return Response({'success': False, 'error_msg': '密码长度至少为6个字符'}, status=400)
    if User.objects.filter(username=username).exists():
        return Response({'success': False, 'error_msg': '用户名已存在'}, status=400)
    if not classroom:
        return Response({'success': False, 'error_msg': '请选择或填写班级'}, status=400)

    u = User.objects.create(
        username=username,
        password=make_password(password),
        role=role,
        department='',
        classroom=classroom
    )
    try:
        u.student_no = username
        u.save(update_fields=['student_no'])
    except Exception:
        pass
    return Response({'success': True, 'message': '注册成功', 'username': u.username, 'role': u.role, 'original_role_request': role_req})


try:
    api_register.throttle_scope = 'register'
except Exception:
    pass


def _parse_bool(val):
    s = str(val).strip().lower()
    if s in {'1','true','y','yes','on','已审核','审核','是'}:
        return True
    if s in {'0','false','n','no','off','未审核','否',''}:
        return False
    try:
        return bool(val)
    except Exception:
        return False


@api_view(["POST"])  # 题目导入（大小与扩展名校验）
def api_question_import(request):
    # 仅老师/管理员可导入
    _, err = require_login(request, roles=['老师','管理员'])
    if err:
        return err
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return Response({"success": False, "error_msg": "未上传有效的 Excel 文件"}, status=400)

    # 大小限制
    try:
        max_bytes = int(getattr(settings, 'EXCEL_MAX_BYTES', 8 * 1024 * 1024))
    except Exception:
        max_bytes = 8 * 1024 * 1024
    size = getattr(excel_file, 'size', None)
    if size is not None and size > max_bytes:
        return Response({"success": False, "error_msg": "Excel 文件过大，限制为 8MB"}, status=400)

    # 扩展名/类型检查（以扩展名为准）
    name = getattr(excel_file, 'name', '') or ''
    if not name.lower().endswith('.xlsx'):
        return Response({"success": False, "error_msg": "仅支持 .xlsx 格式（请用 Office 2007+ 工作簿）"}, status=400)

    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception:
        return Response({"success": False, "error_msg": "Excel 解析失败，请检查模板与内容"}, status=400)
    sheet = wb.active

    # 解析表头：核心列存在即可采用按名导入；否则回退旧模板顺序
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), None)
    header = [str(c.value).strip() if (c and c.value is not None) else '' for c in (header_cells or [])]
    name2idx = {h: i for i, h in enumerate(header) if h}

    def find_idx(*names):
        for n in names:
            if n in name2idx:
                return name2idx[n]
        return None

    # 同义列名
    idx_score   = find_idx('分数','score')
    idx_content = find_idx('内容','题干','content')
    idx_A       = find_idx('A选项','A','A 选项')
    idx_B       = find_idx('B选项','B','B 选项')
    idx_C       = find_idx('C选项','C','C 选项')
    idx_D       = find_idx('D选项','D','D 选项')
    idx_answer  = find_idx('答案','answer')
    idx_analysis= find_idx('解析','analysis')
    idx_qtype   = find_idx('题型','类型','question_type')
    idx_batch   = find_idx('批次','batch')
    idx_kp      = find_idx('考纲大类','知识点','knowledge','knowledge_points')
    idx_pk      = find_idx('一级知识点','primary_knowledge','一级')
    idx_reviewed= find_idx('已审核','reviewed','审核','是否审核')  # 新增：是否审核列

    def get_val(row, idx, default=None):
        if idx is None:
            return default
        return row[idx] if 0 <= idx < len(row) else default

    core_ok = (idx_score is not None and idx_content is not None and idx_answer is not None and idx_qtype is not None)

    # 更新模式（默认开启，可通过 update=false 关闭）
    upd_flag = request.data.get('update') or request.query_params.get('update')
    update_mode = True if upd_flag is None else str(upd_flag).strip().lower() in ('1', 'true', 'yes', 'y')

    created = 0
    updated = 0
    skipped = 0
    errors = 0
    error_rows = []

    allowed_types = {"单选", "多选", "判断", "操作", "Office操作"}

    with dj_tx.atomic():
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or str(v).strip()=='' for v in row):
                skipped += 1
                continue
            try:
                if core_ok:
                    score_raw       = get_val(row, idx_score)
                    content         = get_val(row, idx_content)
                    A               = get_val(row, idx_A)
                    B               = get_val(row, idx_B)
                    C               = get_val(row, idx_C)
                    D               = get_val(row, idx_D)
                    answer          = get_val(row, idx_answer)
                    analysis        = get_val(row, idx_analysis)
                    q_type          = get_val(row, idx_qtype)
                    batch           = get_val(row, idx_batch)
                    knowledge_points= get_val(row, idx_kp, '')
                    primary_knowledge = get_val(row, idx_pk, '')
                    reviewed_raw    = get_val(row, idx_reviewed, None)
                else:
                    # 旧模板列序：score, content, A, B, C, D, answer, analysis, q_type, batch, [kp], [pk]
                    score_raw, content, A, B, C, D, answer, analysis, q_type, batch, *rest = row + (tuple() if isinstance(row, tuple) else ())
                    knowledge_points = (rest[0] or '') if len(rest) >= 1 else ''
                    primary_knowledge = (rest[1] or '') if len(rest) >= 2 else ''
                    reviewed_raw = (rest[2] if len(rest) >= 3 else None)
            except Exception:
                errors += 1
                error_rows.append(idx)
                continue

            # 规范化
            try:
                content = (content or '').strip()
                q_type = (q_type or '').strip()
                answer = (answer or '').strip()
                analysis = (analysis or '').strip()
                A = (A or '').strip() if A else ''
                B = (B or '').strip() if B else ''
                C = (C or '').strip() if C else ''
                D = (D or '').strip() if D else ''
                knowledge_points = (knowledge_points or '').strip()
                primary_knowledge = (primary_knowledge or '').strip()
                reviewed_flag = _parse_bool(reviewed_raw) if reviewed_raw is not None else False
                # 分数：必填且需可转 int
                if score_raw in (None, ''):
                    errors += 1
                    error_rows.append(idx)
                    continue
                try:
                    score = int(score_raw)
                except Exception:
                    try:
                        score = int(float(score_raw))
                    except Exception:
                        errors += 1
                        error_rows.append(idx)
                        continue
                # 批次：允许为空
                if batch is None or str(batch).strip() == '':
                    batch_i = None
                else:
                    try:
                        batch_i = int(batch)
                    except Exception:
                        try:
                            batch_i = int(float(batch))
                        except Exception:
                            batch_i = None
            except Exception:
                errors += 1
                error_rows.append(idx)
                continue

            # 校验（其他字段可为空）
            if not q_type or not content or not answer:
                errors += 1
                error_rows.append(idx)
                continue
            if allowed_types and q_type not in allowed_types:
                errors += 1
                error_rows.append(idx)
                continue
            # 单选/多选：四个选项必须全部非空
            if q_type in {"单选","多选"}:
                if not (A and B and C and D):
                    errors += 1
                    error_rows.append(idx)
                    continue
                import re as _re
                # 规范化答案
                if q_type == '单选':
                    ans_up = str(answer).strip().upper()
                    if ans_up not in {'A','B','C','D'}:
                        errors += 1
                        error_rows.append(idx)
                        continue
                    answer = ans_up
                else:  # 多选
                    letters = _re.findall(r'[A-D]', str(answer).upper())
                    uniq = sorted(set(x for x in letters if x in {'A','B','C','D'}))
                    if len(uniq) < 2:
                        errors += 1
                        error_rows.append(idx)
                        continue
                    answer = ''.join(uniq)

            # 更新或新增
            existing = None
            if update_mode and batch_i is not None:
                try:
                    existing = Question.objects.filter(batch=batch_i, content=content).first()
                except Exception:
                    existing = None
            try:
                if existing:
                    mapping = {
                        'question_type': q_type,
                        'score': score,
                        'A_answer': A,
                        'B_answer': B,
                        'C_answer': C,
                        'D_answer': D,
                        'answer': answer,
                        'analysis': analysis,
                        'knowledge_points': knowledge_points,
                        'primary_knowledge': primary_knowledge,
                    }
                    changed = False
                    for k, v in mapping.items():
                        if getattr(existing, k) != v:
                            setattr(existing, k, v)
                            changed = True
                    # 新增：如果提供了 reviewed，则同步更新（不提供则不动原值）
                    if reviewed_raw is not None:
                        new_flag = bool(reviewed_flag)
                        if bool(getattr(existing, 'reviewed', False)) != new_flag:
                            existing.reviewed = new_flag
                            changed = True
                    if changed:
                        # 如包含 reviewed，也一并写入
                        if reviewed_raw is not None:
                            existing.save(update_fields=list(mapping.keys()) + ['reviewed'])
                        else:
                            existing.save(update_fields=list(mapping.keys()))
                        updated += 1
                    else:
                        skipped += 1
                else:
                    Question.objects.create(
                        question_type=q_type,
                        score=score,
                        content=content,
                        A_answer=A, B_answer=B, C_answer=C, D_answer=D,
                        answer=answer,
                        batch=batch_i,
                        analysis=analysis,
                        knowledge_points=knowledge_points or '',
                        primary_knowledge=primary_knowledge or '',
                        reviewed=bool(reviewed_flag),  # 新增：保存审核标记
                    )
                    created += 1
            except Exception:
                errors += 1
                error_rows.append(idx)
                continue

    return Response({
        'success': True,
        'message': f"题目导入完成: 新增{created}条, 更新{updated}条, 跳过{skipped}条, 错误{errors}条",
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors,
        'error_rows': error_rows[:50],
        'update_mode': update_mode,
        'header_mode': 'by_name' if core_ok else 'legacy_order',
    })


@api_view(["GET"])  # 已发布考试（选择用）
def api_exam_select(request):
    exams = Exam.objects.filter(is_published=True)
    exam_list = [
        {'id': e.exam_id, 'title': e.title, 'start_time': fmt_dt(e.start_time), 'end_time': fmt_dt(e.end_time), 'duration': e.duration}
        for e in exams]
    return Response({'success': True, 'exams': exam_list})


def calculate_score_and_full_score(exam, answers, files):
    batch_ids = [eq.batch for eq in ExamQuestion.objects.filter(exam=exam)]
    questions = list(Question.objects.filter(batch__in=batch_ids))
    full_score = sum(int(getattr(q, 'score', 0) or 0) for q in questions)

    amap = {}
    if isinstance(answers, dict):
        for k, v in answers.items():
            amap[str(k)] = v

    total = 0
    answer_details = []
    op_details = []

    fmap = {}
    try:
        for k in files.keys():
            f = files.get(k)
            m = re.findall(r"(\d+)", str(k))
            if m:
                fmap[int(m[-1])] = f
    except Exception:
        pass

    for q in questions:
        qtype = '操作' if q.question_type in ('操作', 'Office操作') else q.question_type
        qscore = int(getattr(q, 'score', 0) or 0)
        if qtype in ('单选', '多选', '判断'):
            ua = amap.get(str(q.question_id))
            ok = False
            ca = str(q.answer or '')
            if qtype == '单选':
                ok = str(ua).strip().upper() == ca.strip().upper()
            elif qtype == '多选':
                if isinstance(ua, list):
                    uset = set([str(x).strip().upper() for x in ua if str(x).strip()])
                else:
                    letters = re.findall(r'[A-D]', str(ua).upper())
                    uset = set([x.strip().upper() for x in letters])
                cset = set([x for x in re.findall(r'[A-D]', ca.upper())])
                ok = (uset == cset)
            else:  # 判断
                ok = norm_tf(ua) == norm_tf(ca)
            if ok:
                total += qscore
            answer_details.append({'question': q, 'user_answer': ua, 'is_correct': bool(ok), 'type': qtype})
        else:  # 操作题
            fobj = fmap.get(int(q.question_id))
            if not fobj:
                op_details.append({'question_id': q.question_id, 'score': 0, 'total': qscore, 'msg': '未上传文件', 'file_name': None})
                continue
            tmp_path = None
            try:
                name = getattr(fobj, 'name', '') or ''
                _, ext = os.path.splitext(name)
                suffix = (ext or '').lower() or '.bin'
                fd, tmp_path = _tmp.mkstemp(suffix=suffix)
                os.close(fd)
                with open(tmp_path, 'wb') as wf:
                    for chunk in (fobj.chunks() if hasattr(fobj, 'chunks') else [fobj.read()]):
                        if chunk:
                            wf.write(chunk)
                s, t, m = 0, qscore, '不支持的文件类型'
                if suffix in ('.docx', '.doc'):
                    try:
                        s, t, m = score_word(tmp_path)
                    except Exception as e:
                        s, t, m = 0, qscore, f'评分失败: {e}'
                elif suffix in ('.xlsx', '.xls'):
                    try:
                        s, t, m = score_excel(tmp_path)
                    except Exception as e:
                        s, t, m = 0, qscore, f'评分失败: {e}'
                elif suffix in ('.pptx', '.ppt'):
                    try:
                        s, t, m = score_ppt(tmp_path)
                    except Exception as e:
                        s, t, m = 0, qscore, f'评分失败: {e}'
                t = t if isinstance(t, int) and t > 0 else qscore
                try:
                    scaled = int(round(float(s) / float(t) * qscore)) if t > 0 else 0
                except Exception:
                    scaled = 0
                gain = max(0, min(qscore, scaled))
                total += gain
                op_details.append({'question_id': q.question_id, 'score': gain, 'total': qscore, 'msg': m, 'file_name': getattr(fobj, 'name', None)})
            finally:
                try:
                    if tmp_path and os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except Exception:
                    pass

    return total, full_score, answer_details, op_details


@api_view(["POST"])  # 提交考试
def api_submit_exam(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return Response({'success': False, 'error_msg': '请先登录'}, status=401)
    try:
        user = User.objects.get(user_id=user_id)
    except User.DoesNotExist:
        return Response({'success': False, 'error_msg': '用户不存在'}, status=401)

    exam_id = (
        request.data.get('exam_id') or request.data.get('id') or request.data.get('examId') or
        request.POST.get('exam_id') or request.POST.get('id') or request.POST.get('examId')
    )
    if not exam_id:
        return Response({'success': False, 'error_msg': '缺少考试ID'})
    try:
        exam = Exam.objects.get(exam_id=int(exam_id))
    except Exception:
        return Response({'success': False, 'error_msg': '考试不存在'})

    if ExamRecord.objects.filter(exam=exam, user=user).exists():
        return Response({'success': False, 'error_msg': '该考试已提交，不能再次考试'})

    from django.conf import settings
    try:
        ttl = int(getattr(settings, 'FACE_SIGNIN_TTL_MINUTES', 120) or 120)
    except Exception:
        ttl = 120
    since = timezone.now() - timedelta(minutes=ttl)
    try:
        from ..models import ExamSignIn as _Sign
    except Exception:
        _Sign = None
    if _Sign is not None:
        ok = _Sign.objects.filter(exam=exam, user=user, success=True, created_at__gte=since).exists()
        if not ok:
            return Response({'success': False, 'error_msg': '未在人脸识别通过后参与本场考试或已过期，请先在本场考试签到'}, status=403)

    now = timezone.now()
    if not exam.is_published or exam.start_time > now or exam.end_time < now:
        return Response({'success': False, 'error_msg': '当前不在考试有效时间内，无法提交'})

    answers = request.data.get('answers')
    if isinstance(answers, str):
        try:
            import json
            answers = json.loads(answers)
        except Exception:
            answers = None
    if not isinstance(answers, dict):
        answers = {}
        for k, v in request.data.items():
            if str(k).startswith('answer_'):
                qid = k.replace('answer_', '')
                answers[qid] = v

    files = request.FILES

    score, full_score, answer_details, op_details = calculate_score_and_full_score(exam, answers, files)

    rec = ExamRecord.objects.create(
        user=user,
        exam=exam,
        score=score,
        submit_time=now,
        start_time=now,
        end_time=now,
    )

    bulk = []
    for it in (answer_details or []):
        if it.get('type') not in ('单选', '多选', '判断'):
            continue
        bulk.append(AnswerRecord(
            record=rec,
            question=it['question'],
            user_answer=str(it.get('user_answer', '')),
            is_correct=bool(it.get('is_correct', False)),
        ))
    if bulk:
        AnswerRecord.objects.bulk_create(bulk, ignore_conflicts=True)

    try:
        op_bulk = []
        for it in (op_details or []):
            try:
                qobj = Question.objects.get(question_id=int(it.get('question_id')))
            except Exception:
                qobj = None
            if not qobj:
                # 跳过未找到题目的明细，避免非空外键错误
                continue
            op_bulk.append(OperationDetail(
                record=rec,
                question=qobj,
                score=it.get('score'),
                total=it.get('total'),
                msg=it.get('msg'),
                file_name=it.get('file_name')
            ))
        if op_bulk:
            OperationDetail.objects.bulk_create(op_bulk, ignore_conflicts=True)
    except Exception:
        pass

    return Response({'success': True, 'score': score, 'full_score': full_score, 'record_id': rec.record_id})


@api_view(["GET"])  # 成绩详情
def api_score_detail(request):
    uid = request.session.get('user_id')
    if not uid:
        return Response({'success': False, 'error_msg': '未登录'}, status=401)
    record_id = request.query_params.get('record_id') or request.query_params.get('id')
    if not record_id:
        return Response({'success': False, 'error_msg': '缺少记录ID'})
    try:
        rec = ExamRecord.objects.select_related('exam', 'user').get(record_id=int(record_id), user__user_id=uid)
    except Exception:
        return Response({'success': False, 'error_msg': '记录不存在或无权查看'})

    items = []
    for ar in AnswerRecord.objects.select_related('question').filter(record=rec):
        q = ar.question
        items.append({
            'question_id': getattr(q, 'question_id', None),
            'type': getattr(q, 'question_type', ''),
            'content': getattr(q, 'content', ''),
            'user_answer': ar.user_answer,
            'correct_answer': getattr(q, 'answer', ''),
            'is_correct': bool(ar.is_correct),
            'q_score': (getattr(q, 'score', 0) or 0) if bool(ar.is_correct) else 0,
            'q_total': getattr(q, 'score', 0) or 0,
            'analysis': getattr(q, 'analysis', '') or ''
        })
    op_details = list(OperationDetail.objects.filter(record=rec))
    if op_details:
        q_map = {q.question_id: q for q in Question.objects.filter(question_id__in=[d.question_id for d in op_details if d.question_id])}
        for d in op_details:
            q = q_map.get(getattr(d, 'question_id', None)) if q_map else None
            items.append({
                'question_id': getattr(d.question, 'question_id', None),
                'type': '操作',
                'content': getattr(q, 'content', ''),
                'user_answer': '[file]',
                'correct_answer': '',
                'is_correct': None,
                'op_score': d.score,
                'op_total': d.total,
                'op_msg': d.msg,
                'file_name': d.file_name,
                'q_score': d.score,
                'q_total': d.total,
                'analysis': getattr(q, 'analysis', '') or ''
            })

    batch_ids = [eq.batch for eq in ExamQuestion.objects.filter(exam=rec.exam)]
    full_score = sum(getattr(q, 'score', 1) for q in Question.objects.filter(batch__in=batch_ids))

    return Response({'success': True, 'record_id': rec.record_id, 'exam_title': rec.exam.title,
                     'score': rec.score,
                     'full_score': full_score,
                     'submit_time': fmt_dt(rec.submit_time), 'details': items})


@api_view(["GET"])  # 题目评论列表（需已作答）
def api_q_comments(request):
    user, err = require_login(request)
    if err:
        return err
    try:
        qid = int(request.query_params.get('question_id') or 0)
    except Exception:
        qid = 0
    if not qid:
        return Response({'success': False, 'error_msg': '缺少题目ID'})
    # 校验是否已作答（提交判题）
    if not PracticeAnswer.objects.filter(user_id=user.user_id, question_id=qid).exists():
        return Response({'success': False, 'error_msg': '请先提交本题后再查看评论'}, status=403)
    qs = QuestionComment.objects.filter(question_id=qid).select_related('user').annotate(like_count=Count('likes')).order_by('-like_count', '-id')[:100]
    ids = [c.id for c in qs]
    liked_ids = set(QuestionCommentLike.objects.filter(comment_id__in=ids, user_id=user.user_id).values_list('comment_id', flat=True)) if ids else set()
    items = []
    for c in qs:
        u = c.user
        uname = (getattr(u, 'nickname', '') or '').strip() or (getattr(u, 'username', '') or '').strip() or '匿名用户'
        items.append({
            'id': c.id,
            'user_id': getattr(u, 'user_id', None),
            'username': uname,
            'content': c.content,
            'created_at': fmt_dt(c.created_at),
            'likes': int(getattr(c, 'like_count', 0) or 0),
            'liked_by_me': c.id in liked_ids,
            'can_delete': c.user_id == user.user_id
        })
    return Response({'success': True, 'items': items})


@api_view(["POST"])  # 发表题目评论（需已作答）
def api_q_comments_create(request):
    user, err = require_login(request)
    if err:
        return err
    try:
        qid = int(request.data.get('question_id') or 0)
    except Exception:
        qid = 0
    content = (request.data.get('content') or '').strip()
    if not qid:
        return Response({'success': False, 'error_msg': '缺少题目ID'})
    if not content:
        return Response({'success': False, 'error_msg': '内容不能为空'})
    # 基本长度限制
    if not (2 <= len(content) <= 500):
        return Response({'success': False, 'error_msg': '评论长度需在 2-500 字之间'})
    # 简单敏感词过滤（可扩展为正则/词库）
    bad_words = ['傻逼','sb','草你','操你','妈的','垃圾','fuck','shit','艹','滚','狗屎']
    lower = content.lower()
    for w in bad_words:
        if w in lower:
            return Response({'success': False, 'error_msg': '包含不当用语，请文明发言'})
    if not PracticeAnswer.objects.filter(user_id=user.user_id, question_id=qid).exists():
        return Response({'success': False, 'error_msg': '请先提交本题后再发表评论'}, status=403)
    # 节流：同一用户发言间隔至少 10 秒；同一题同样内容 5 分钟内仅一次；10 分钟内最多 10 条
    now = timezone.now()
    last = QuestionComment.objects.filter(user_id=user.user_id).order_by('-id').first()
    if last and (now - last.created_at).total_seconds() < 10:
        return Response({'success': False, 'error_msg': '发言过于频繁，请稍后再试'})
    from datetime import timedelta as _td
    dup = QuestionComment.objects.filter(user_id=user.user_id, question_id=qid, content=content, created_at__gte=now - _td(minutes=5)).exists()
    if dup:
        return Response({'success': False, 'error_msg': '请勿重复提交相同内容'})
    recent_count = QuestionComment.objects.filter(user_id=user.user_id, created_at__gte=now - _td(minutes=10)).count()
    if recent_count >= 10:
        return Response({'success': False, 'error_msg': '发言过多，请稍后再试'})
    try:
        QuestionComment.objects.create(user=user, question_id=qid, content=content)
        return Response({'success': True})
    except Exception as e:
        return Response({'success': False, 'error_msg': f'保存失败: {e}'})


@api_view(["POST"])  # 删除题目评论（本人/老师/管理员）
def api_q_comments_delete(request):
    user, err = require_login(request)
    if err:
        return err
    try:
        cid = int(request.data.get('id') or 0)
    except Exception:
        cid = 0
    if not cid:
        return Response({'success': False, 'error_msg': '缺少评论ID'})
    c = QuestionComment.objects.filter(id=cid).select_related('user').first()
    if not c:
        return Response({'success': False, 'error_msg': '评论不存在'})
    # 权限判断：本人或老师/管理员
    role = getattr(user, 'role', '')
    if not (c.user_id == user.user_id or role in ('老师','管理员')):
        return Response({'success': False, 'error_msg': '无权删除'}, status=403)
    c.delete()
    return Response({'success': True})


@api_view(["POST"])  # 评论点赞（需已作答）
def api_q_comment_like(request):
    user, err = require_login(request)
    if err:
        return err
    try:
        cid = int(request.data.get('id') or 0)
    except Exception:
        cid = 0
    if not cid:
        return Response({'success': False, 'error_msg': '缺少评论ID'})
    c = QuestionComment.objects.filter(id=cid).first()
    if not c:
        return Response({'success': False, 'error_msg': '评论不存在'})
    # 需先作答该题
    if not PracticeAnswer.objects.filter(user_id=user.user_id, question_id=c.question_id).exists():
        return Response({'success': False, 'error_msg': '请先提交本题后再点赞'}, status=403)
    like, created = QuestionCommentLike.objects.get_or_create(comment=c, user=user)
    # 返回最新计数
    cnt = QuestionCommentLike.objects.filter(comment=c).count()
    return Response({'success': True, 'liked': True, 'likes': cnt})


@api_view(["POST"])  # 取消点赞（需已作答）
def api_q_comment_unlike(request):
    user, err = require_login(request)
    if err:
        return err
    try:
        cid = int(request.data.get('id') or 0)
    except Exception:
        cid = 0
    if not cid:
        return Response({'success': False, 'error_msg': '缺少评论ID'})
    c = QuestionComment.objects.filter(id=cid).first()
    if not c:
        return Response({'success': False, 'error_msg': '评论不存在'})
    # 需先作答该题
    if not PracticeAnswer.objects.filter(user_id=user.user_id, question_id=c.question_id).exists():
        return Response({'success': False, 'error_msg': '请先提交本题后再操作'}, status=403)
    QuestionCommentLike.objects.filter(comment=c, user=user).delete()
    cnt = QuestionCommentLike.objects.filter(comment=c).count()
    return Response({'success': True, 'liked': False, 'likes': cnt})
